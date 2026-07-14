from __future__ import annotations

import json
import shutil
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
INPUT_XLSX = ROOT / "data" / "bottle_mail" / "TeaMerry_Drift_Bottle_Mail_Master_v02.xlsx"
OUTPUT_JSON = ROOT / "data" / "export" / "drift_bottle_messages.json"
DOCS_JSON = ROOT / "docs" / "data" / "export" / "drift_bottle_messages.json"

HANDWRITING_IDS = {"quiet", "round", "careful", "faded", "child"}

COLUMN_ALIASES = {
    "id": {"ID", "メールID", "id"},
    "category": {"カテゴリ", "カテゴリー", "category"},
    "displayName": {"表示名", "表示名（空欄＝おさんぽさん）", "表示名(空欄=おさんぽさん)"},
    "message": {"ボトル本文", "本文", "message", "text"},
    "handwritingTemplate": {"筆跡テンプレート", "筆跡", "handwritingTemplate"},
    "hotNewsHistory": {"ほっこり表示履歴", "ほっこり履歴", "今日のほっこり"},
    "note": {"備考", "note"},
}


def normalize_text(value) -> str:
    return str(value).strip() if value is not None else ""


def normalize_header(value) -> str:
    return normalize_text(value).replace("\n", "").replace(" ", "").replace("　", "").replace("＝", "=")


def find_header_row(sheet) -> tuple[int, dict[str, int]]:
    normalized_aliases = {
        key: {normalize_header(alias) for alias in aliases}
        for key, aliases in COLUMN_ALIASES.items()
    }

    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        headers = [normalize_header(value) for value in row]
        mapping: dict[str, int] = {}

        for index, header in enumerate(headers):
          for key, aliases in normalized_aliases.items():
              if header in aliases and key not in mapping:
                  mapping[key] = index

        if {"id", "category", "displayName", "message", "handwritingTemplate"}.issubset(mapping):
            return row_number, mapping

    raise ValueError("Ver.2ボトルメールのヘッダー行が見つかりません")


def normalize_handwriting(value) -> str:
    text = normalize_text(value)
    if not text:
        return "quiet"

    template_id = text.split("（", 1)[0].split("(", 1)[0].strip().lower()
    return template_id if template_id in HANDWRITING_IDS else "quiet"


def get_cell(row, mapping: dict[str, int], key: str) -> str:
    index = mapping.get(key)
    return normalize_text(row[index]) if index is not None and index < len(row) else ""


def build_messages(workbook):
    sheet = workbook["本文"] if "本文" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    header_row, mapping = find_header_row(sheet)
    messages: list[dict] = []
    errors: list[str] = []
    duplicate_ids: list[str] = []
    seen_ids: set[str] = set()
    counts = Counter()

    for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        raw_values = [normalize_text(value) for value in row]
        if not any(raw_values):
            continue

        message_id = get_cell(row, mapping, "id")
        message_text = get_cell(row, mapping, "message")

        if not message_id:
            counts["missing_id"] += 1
            continue

        if not message_text:
            counts["empty_message"] += 1
            continue

        if message_id in seen_ids:
            duplicate_ids.append(message_id)
            continue
        seen_ids.add(message_id)

        message_length = len(message_text)
        if message_length > 100:
            counts["over_100"] += 1
            errors.append(f"{message_id}: 本文が100文字を超えています ({message_length}文字)")
            continue

        display_name = get_cell(row, mapping, "displayName")
        if not display_name:
            display_name = "おさんぽさん"
            counts["filled_display_name"] += 1

        handwriting_template = normalize_handwriting(get_cell(row, mapping, "handwritingTemplate"))

        message = {
            "id": message_id.strip(),
            "category": get_cell(row, mapping, "category"),
            "displayName": display_name,
            "message": message_text,
            "text": message_text,
            "handwritingTemplate": handwriting_template,
            "enabled": True,
        }

        hot_news_history = get_cell(row, mapping, "hotNewsHistory")
        note = get_cell(row, mapping, "note")
        if hot_news_history:
            message["hotNewsHistory"] = hot_news_history
        if note:
            message["note"] = note

        messages.append(message)

    if duplicate_ids:
        counts["duplicate_id"] = len(duplicate_ids)
        errors.append("ID重複: " + ", ".join(sorted(set(duplicate_ids))))

    return messages, counts, errors


def write_json(messages: list[dict]) -> None:
    payload = {
        "version": "2.0",
        "source": INPUT_XLSX.name,
        "updatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "messages": messages,
    }

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    DOCS_JSON.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(OUTPUT_JSON, DOCS_JSON)


def print_report(messages: list[dict], counts: Counter) -> None:
    print(f"Exported: {len(messages)}")
    print(f"Display name filled: {counts['filled_display_name']}")
    print(f"Missing ID rows: {counts['missing_id']}")
    print(f"Empty message rows: {counts['empty_message']}")
    print(f"Duplicate ID rows: {counts['duplicate_id']}")
    print(f"Over 100 chars rows: {counts['over_100']}")
    print("Handwriting templates:")
    for key, value in sorted(Counter(message["handwritingTemplate"] for message in messages).items()):
        print(f"  {key}: {value}")
    print("Categories:")
    for key, value in sorted(Counter(message["category"] for message in messages).items()):
        print(f"  {key}: {value}")
    print(f"Output: {OUTPUT_JSON}")
    print(f"Docs copy: {DOCS_JSON}")


def main() -> int:
    if not INPUT_XLSX.exists():
        print(f"入力Excelが存在しません: {INPUT_XLSX}", file=sys.stderr)
        return 1

    workbook = load_workbook(INPUT_XLSX, read_only=False, data_only=True)
    try:
        messages, counts, errors = build_messages(workbook)
    finally:
        workbook.close()

    if errors:
        for error in errors:
            print(error, file=sys.stderr)
        print_report(messages, counts)
        return 1

    write_json(messages)
    print_report(messages, counts)
    return 0


if __name__ == "__main__":
    sys.exit(main())
