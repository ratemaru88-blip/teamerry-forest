from __future__ import annotations

import json
import shutil
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
INPUT_XLSX = ROOT / "data" / "dialogue" / "TeaMerry_Dialogue_Spreadsheet_Template.xlsx"
OUTPUT_JSON = ROOT / "data" / "export" / "dialogue.json"
OUTPUT_REPORT = ROOT / "data" / "export" / "dialogue_export_report.md"
DOCS_JSON = ROOT / "docs" / "data" / "export" / "dialogue.json"

TARGET_SHEETS = ["Mint", "Lil", "Elder", "Maroud", "Forest", "Other"]
EXPECTED_HEADERS = [
    "セリフID",
    "表示種別",
    "キャラクター",
    "場所",
    "章",
    "条件",
    "トーン",
    "優先度",
    "セリフ本文",
    "有効",
    "備考",
]

TYPE_ALIASES = {
    "forest": "forest_whisper",
}


def normalize_text(value) -> str:
    return str(value).strip() if value is not None else ""


def normalize_type(value) -> str:
    text = normalize_text(value)
    return TYPE_ALIASES.get(text, text)


def parse_conditions(value) -> list[str]:
    text = normalize_text(value)
    if not text:
        return []
    return [item.strip() for item in text.split(",") if item.strip()]


def parse_priority(value, warnings: list[str], sheet_name: str, row_number: int) -> int | float:
    text = normalize_text(value)
    if not text:
        warnings.append(f"{sheet_name} row {row_number}: 優先度空欄のため 50 を使用")
        return 50

    try:
        number = float(text)
    except ValueError:
        warnings.append(f"{sheet_name} row {row_number}: 優先度不正 '{text}' のため 50 を使用")
        return 50

    return int(number) if number.is_integer() else number


def is_repeated_header(values: list[str]) -> bool:
    return values == EXPECTED_HEADERS


def validate_workbook(workbook) -> list[str]:
    errors: list[str] = []

    for sheet_name in TARGET_SHEETS:
        if sheet_name not in workbook.sheetnames:
            errors.append(f"対象シートが存在しません: {sheet_name}")
            continue

        sheet = workbook[sheet_name]
        headers = [normalize_text(sheet.cell(row=1, column=index + 1).value) for index in range(11)]
        if headers != EXPECTED_HEADERS:
            errors.append(f"{sheet_name}: 11列ヘッダーが一致しません")

    return errors


def build_dialogues(workbook):
    dialogues: list[dict] = []
    warnings: list[str] = []
    sheet_counts: Counter[str] = Counter()
    type_counts: Counter[str] = Counter()
    excluded_counts: dict[str, Counter[str]] = defaultdict(Counter)
    seen_ids: dict[str, str] = {}
    duplicate_ids: list[str] = []

    for sheet_name in TARGET_SHEETS:
        sheet = workbook[sheet_name]

        for row_number, row in enumerate(sheet.iter_rows(min_row=2, max_col=11, values_only=True), start=2):
            values = [normalize_text(value) for value in row]

            if not any(values):
                excluded_counts[sheet_name]["空行"] += 1
                continue

            if is_repeated_header(values):
                excluded_counts[sheet_name]["再掲ヘッダー行"] += 1
                continue

            dialogue_id = values[0]
            dialogue_text = values[8]
            enabled_text = values[9].upper()

            if not dialogue_id:
                excluded_counts[sheet_name]["セリフIDなし"] += 1
                continue

            if not dialogue_text:
                excluded_counts[sheet_name]["セリフ本文なし"] += 1
                warnings.append(f"{sheet_name} row {row_number}: セリフ本文空欄のため除外")
                continue

            if enabled_text == "OFF":
                excluded_counts[sheet_name]["有効OFF"] += 1
                continue

            if enabled_text != "ON":
                excluded_counts[sheet_name]["有効値不正"] += 1
                warnings.append(f"{sheet_name} row {row_number}: 有効値不正 '{values[9]}' のため除外")
                continue

            if dialogue_id in seen_ids:
                duplicate_ids.append(dialogue_id)
            else:
                seen_ids[dialogue_id] = f"{sheet_name} row {row_number}"

            dialogue_type = normalize_type(values[1])
            if not dialogue_type:
                warnings.append(f"{sheet_name} row {row_number}: 表示種別空欄")

            dialogue = {
                "id": dialogue_id,
                "type": dialogue_type,
                "character": values[2],
                "place": values[3],
                "section": values[4],
                "conditions": parse_conditions(values[5]),
                "tone": values[6],
                "priority": parse_priority(values[7], warnings, sheet_name, row_number),
                "text": dialogue_text,
                "enabled": True,
                "note": values[10],
            }

            dialogues.append(dialogue)
            sheet_counts[sheet_name] += 1
            type_counts[dialogue_type] += 1

    return dialogues, warnings, sheet_counts, type_counts, excluded_counts, duplicate_ids


def write_json(dialogues: list[dict]) -> None:
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)

    payload = {
        "version": "1.0",
        "source": "TeaMerry_Dialogue_Spreadsheet_Template.xlsx",
        "updatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "dialogues": dialogues,
    }

    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    DOCS_JSON.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(OUTPUT_JSON, DOCS_JSON)


def write_report(
    success: bool,
    dialogues: list[dict],
    warnings: list[str],
    errors: list[str],
    sheet_counts: Counter[str],
    type_counts: Counter[str],
    excluded_counts: dict[str, Counter[str]],
) -> None:
    OUTPUT_REPORT.parent.mkdir(parents=True, exist_ok=True)

    lines = [
        "# TeaMerry Dialogue Export Report",
        "",
        f"- Export日時: {datetime.now().astimezone().isoformat(timespec='seconds')}",
        f"- 入力ファイル: `{INPUT_XLSX.relative_to(ROOT)}`",
        f"- 出力ファイル: `{OUTPUT_JSON.relative_to(ROOT)}`",
        f"- docs公開用コピー: `{DOCS_JSON.relative_to(ROOT)}`",
        f"- 総出力件数: {len(dialogues)}",
        f"- Export結果: {'成功' if success else '失敗'}",
        "",
        "## シート別件数",
        "",
    ]

    for sheet_name in TARGET_SHEETS:
        lines.append(f"- {sheet_name}: {sheet_counts.get(sheet_name, 0)}")

    lines.extend(["", "## 表示種別別件数", ""])
    if type_counts:
        for type_name, count in sorted(type_counts.items()):
            lines.append(f"- {type_name}: {count}")
    else:
        lines.append("- なし")

    lines.extend(["", "## 除外件数", ""])
    for sheet_name in TARGET_SHEETS:
        excluded = excluded_counts.get(sheet_name, Counter())
        if excluded:
            detail = ", ".join(f"{name}: {count}" for name, count in sorted(excluded.items()))
            lines.append(f"- {sheet_name}: {detail}")
        else:
            lines.append(f"- {sheet_name}: なし")

    lines.extend(["", "## 警告一覧", ""])
    if warnings:
        lines.extend(f"- {warning}" for warning in warnings)
    else:
        lines.append("- なし")

    lines.extend(["", "## エラー一覧", ""])
    if errors:
        lines.extend(f"- {error}" for error in errors)
    else:
        lines.append("- なし")

    OUTPUT_REPORT.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    errors: list[str] = []

    if not INPUT_XLSX.exists():
        errors.append(f"入力Excelが存在しません: {INPUT_XLSX}")
        write_report(False, [], [], errors, Counter(), Counter(), defaultdict(Counter))
        return 1

    workbook = load_workbook(INPUT_XLSX, read_only=True, data_only=True)
    try:
        errors.extend(validate_workbook(workbook))

        if errors:
            write_report(False, [], [], errors, Counter(), Counter(), defaultdict(Counter))
            return 1

        dialogues, warnings, sheet_counts, type_counts, excluded_counts, duplicate_ids = build_dialogues(workbook)

        if duplicate_ids:
            errors.append("セリフID重複: " + ", ".join(sorted(set(duplicate_ids))))
            write_report(False, dialogues, warnings, errors, sheet_counts, type_counts, excluded_counts)
            return 1

        write_json(dialogues)
        write_report(True, dialogues, warnings, errors, sheet_counts, type_counts, excluded_counts)
    finally:
        workbook.close()

    print(f"Exported {len(dialogues)} dialogues to {OUTPUT_JSON}")
    print(f"Copied dialogue JSON to {DOCS_JSON}")
    print(f"Wrote report to {OUTPUT_REPORT}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
